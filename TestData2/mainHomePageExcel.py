import openpyxl

class MainHomePageData_Excel:

    def Main_HomePage_Function_Excel():
        book=openpyxl.load_workbook("pyExcelDemo.xlsx")
        sheet=book.active
        #value=sheet.cell(row=4,column=2).value

        listingof_dict=[]



        for i in range(2,sheet.max_row+1):
            listing={}
            for j in range(2,sheet.max_column+1):
                values=sheet.cell(row=i,column=j).value

                
                dic=sheet.cell(row=1,column=j).value
                listing[dic]=values
                
            listingof_dict.append(listing)


        return listingof_dict


