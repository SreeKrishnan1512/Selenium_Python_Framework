import openpyxl
from itertools import zip_longest

book=openpyxl.load_workbook("C:\\Users\\psaik\\Desktop\\Python_Selenium_Framework2\\TestData2\\pyExcelDemo.xlsx")
sheet=book.active
cell=sheet.cell(row=4,column=2)
print(cell.value)
print(sheet['B3'].value)

#new=sheet.cell(row=4,column=2).value="Dev"
#print(new)
print(sheet.max_row)
print(sheet.max_column)

print("entering into for loop")
print()
for i in range(2,sheet.max_row+1):
    for j in range(2,sheet.max_column+1):
         values=sheet.cell(row=i,column=j).value
         print(values)
    print("\n")

#book.save("C:\\Users\\psaik\\Desktop\\Python_Selenium_Framework2\\TestData2\\pyExcelDemo.xlsx")
#print("Value saved successfully!")

