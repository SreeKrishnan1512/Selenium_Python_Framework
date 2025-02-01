import openpyxl

book=openpyxl.load_workbook("C:\\Users\\psaik\\Desktop\\Python_Selenium_Framework2\\TestData2\\pyExcelDemo.xlsx")
sheet=book.active
#value=sheet.cell(row=4,column=2).value

listingoflist=[]

for j in range (2,sheet.max_column+1):
    values=sheet.cell(row=1,column=j).value
    print(values)